from openpyxl import load_workbook, Workbook

class ExcelEditorConsoleApp:
    DEFAULT_FILE_PATH = "C:/Users/MagicBook Pro/Desktop/Учеба/Программы/sss.xlsx"

    def __init__(self):
        self.file_path = None

    def load_file(self):
        self.file_path = input("Введите путь до Эксель-таблицы: ")
        try:
            workbook = load_workbook(self.file_path)
            sheet = workbook.active
            print(f"Загружен файл: {self.file_path}")
        except FileNotFoundError:
            print(f"Файл не найден по указанному пути: {self.file_path}")
        except Exception as e:
            print(f"Ошибка при загрузке файла: {e}")

    def read_excel(self):
        if not self.file_path:
            print("Файл не был загружен. Используется файл по умолчанию.")
            self.file_path = self.DEFAULT_FILE_PATH

        try:
            workbook = load_workbook(self.file_path)
            sheet = workbook.active
            print(f"Содержимое {self.file_path}:\n")
            for row in sheet.iter_rows(values_only=True):
                print("\t".join(map(str, row)))
        except Exception as e:
            print(f"Ошибка при чтении файла: {e}")

    def write_excel(self):
        if not self.file_path:
            print("Файл не был загружен. Используется файл по умолчанию.")
            self.file_path = self.DEFAULT_FILE_PATH

        try:
            workbook = Workbook()
            sheet = workbook.active
            content_to_write = input(
                "Введите данные для записи (Вводите через запятую, в конце нажмите Enter):\n"
            )
            data = [item.split(',') for item in content_to_write.split('\n')]

            for row in data:
                sheet.append(row)

            save_choice = input("Сохранить как новый файл? (y/n): ").lower()
            if save_choice == 'y':
                new_file_path = input("Введите путь для сохранения новой Эксель-таблицы: ")
                workbook.save(new_file_path)
                print(f"Данные записаны в {new_file_path}")
            else:
                workbook.save(self.file_path)
                print(f"Данные сохранены в существующем файле: {self.file_path}")

        except Exception as e:
            print(f"Ошибка при записи файла: {e}")

    def delete_row(self):
        if self.file_path:
            try:
                workbook = load_workbook(self.file_path)
                sheet = workbook.active
                row_to_delete = int(input("Введите номер строки для удаления: "))
                sheet.delete_rows(row_to_delete)
                workbook.save(self.file_path)
                print(f"Строка {row_to_delete} удалена из файла: {self.file_path}")
            except Exception as e:
                print(f"Ошибка при удалении строки: {e}")
        else:
            print("Файл не был загружен. Загрузите файл перед удалением строки.")

if __name__ == "__main__":
    app = ExcelEditorConsoleApp()

    while True:
        print("\nМеню:")
        print("1. Загрузить файл")
        print("2. Просмотреть файл")
        print("3. Заполнить файл")
        print("4. Удалить строку")
        print("5. Выйти")

        choice = input("Введите свой выбор (1-5): ")

        if choice == "1":
            app.load_file()
        elif choice == "2":
            app.read_excel()
        elif choice == "3":
            app.write_excel()
        elif choice == "4":
            app.delete_row()
        elif choice == "5":
            break
        else:
            print("Неверный выбор. Пожалуйста, введите допустимую опцию.")
