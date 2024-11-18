import socket
import os
from openpyxl import load_workbook

SERVER_HOST = '127.0.0.1'
SERVER_PORT = 5000
BUFFER_SIZE = 1024

def upload_file(client_socket, filename):
    client_socket.sendall(b"READY")
    file_data = b''
    while True:
        data = client_socket.recv(BUFFER_SIZE)
        if data == b'DONE':
            break
        file_data += data
    with open(filename, 'wb') as f:
        f.write(file_data)
    client_socket.sendall(b"File uploaded successfully.")

def download_file(client_socket, filename):
    if os.path.exists(filename):
        client_socket.sendall(b"EXIST")
        wb = load_workbook(filename=filename)
        ws = wb.active
        ws_data = ws.iter_rows(values_only=True)
        for row in ws_data:
            client_socket.sendall(str(row).encode())
        client_socket.sendall(b"DONE")
        wb.close()
    else:
        client_socket.sendall(b"NOT_EXIST")

def modify_file(client_socket, filename, row_number, new_value):
    if os.path.exists(filename):
        client_socket.sendall(b"EXIST")
        wb = load_workbook(filename=filename)
        ws = wb.active
        ws.cell(row=row_number, column=1, value=new_value)
        wb.save(filename)
        client_socket.sendall(b"Changes applied successfully.")
    else:
        client_socket.sendall(b"NOT_EXIST")

def handle_client_connection(client_socket):
    while True:
        request = client_socket.recv(BUFFER_SIZE).decode()
        if not request:
            break

        if request == 'UPLOAD':
            filename = client_socket.recv(BUFFER_SIZE).decode()
            upload_file(client_socket, filename)
        
        elif request == 'DOWNLOAD':
            filename = client_socket.recv(BUFFER_SIZE).decode()
            download_file(client_socket, filename)

        elif request == 'MODIFY':
            filename = client_socket.recv(BUFFER_SIZE).decode()
            row_number = int(client_socket.recv(BUFFER_SIZE).decode())
            new_value = client_socket.recv(BUFFER_SIZE).decode()
            modify_file(client_socket, filename, row_number, new_value)

        else:
            client_socket.sendall(b"Invalid command.")

    client_socket.close()

def main():
    server_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    server_socket.bind((SERVER_HOST, SERVER_PORT))
    server_socket.listen(5)
    print(f"[*] Listening on {SERVER_HOST}:{SERVER_PORT}")

    while True:
        client_socket, _ = server_socket.accept()
        print(f"[*] Accepted connection from {client_socket.getpeername()}")

        handle_client_connection(client_socket)

if __name__ == "__main__":
    main()
